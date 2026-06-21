import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

class UniversalExcelManager:
    def __init__(self, filename='store_bills.xlsx'):
        self.filename = filename
        self.init_excel_file()

    def init_excel_file(self):
        if not os.path.exists(self.filename):
            wb = Workbook()

            # ----- Bills Sheet -----
            ws_bills = wb.active
            ws_bills.title = "Bills"
            headers_bills = [
                "Sr No", "Date", "Customer Name", "Customer ID", "Bill No",
                "Items Count", "Subtotal", "Discount", "Grand Total", "QR Code", "PDF File"
            ]
            for col, header in enumerate(headers_bills, 1):
                cell = ws_bills.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
            col_widths = {'A':8,'B':12,'C':22,'D':15,'E':22,'F':10,'G':15,'H':15,'I':15,'J':20,'K':20}
            for col, width in col_widths.items():
                ws_bills.column_dimensions[col].width = width

            # ----- Profit & Loss Sheet -----
            ws_pl = wb.create_sheet("ProfitLoss")
            headers_pl = ["Bill No","Date","Subtotal","Discount","Grand Total","Profit","Type"]
            for col, header in enumerate(headers_pl, 1):
                cell = ws_pl.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
            col_widths_pl = {'A':22,'B':12,'C':15,'D':15,'E':15,'F':15,'G':10}
            for col, width in col_widths_pl.items():
                ws_pl.column_dimensions[col].width = width

            # ----- Returns Sheet -----
            ws_ret = wb.create_sheet("Returns")
            headers_ret = ["Sr No","Date","Customer Name","Customer ID","Original Bill No",
                           "Return Bill No","Item Name","Qty Returned","Unit Price",
                           "Return Amount","Reason"]
            for col, header in enumerate(headers_ret, 1):
                cell = ws_ret.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            col_widths_ret = {'A':8,'B':12,'C':22,'D':15,'E':22,'F':22,'G':25,'H':12,'I':12,'J':15,'K':25}
            for col, width in col_widths_ret.items():
                ws_ret.column_dimensions[col].width = width

            # ----- Settings Sheet -----
            ws_settings = wb.create_sheet("Settings")
            ws_settings.append(["Setting", "Value"])
            ws_settings.append(["store_name", "My Store"])
            ws_settings.append(["store_tagline", "Quality You Can Trust"])
            ws_settings.append(["store_address", "123 Main Street, City - 000000"])
            ws_settings.append(["store_phone", "+91 98765 43210"])
            ws_settings.append(["store_email", "contact@mystore.com"])
            ws_settings.append(["store_website", "www.mystore.com"])
            ws_settings.append(["currency_symbol", "₹"])
            ws_settings.append(["profit_margin", 30])
            ws_settings.append(["customer_counter", 1000])
            ws_settings.append(["bill_counter", 1])
            ws_settings.append(["return_counter", 1])
            ws_settings.append(["session_timeout_minutes", 10])  # Admin-configurable
            for col in range(1, 3):
                cell = ws_settings.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            ws_settings.column_dimensions['A'].width = 30
            ws_settings.column_dimensions['B'].width = 40

            wb.save(self.filename)
        else:
            # Migrate existing file: add Returns sheet and new settings if missing
            self._migrate_existing_file()

    def _migrate_existing_file(self):
        """Add missing sheets/settings to existing Excel files."""
        try:
            wb = load_workbook(self.filename)
            changed = False

            # Add Returns sheet if missing
            if "Returns" not in wb.sheetnames:
                ws_ret = wb.create_sheet("Returns")
                headers_ret = ["Sr No","Date","Customer Name","Customer ID","Original Bill No",
                               "Return Bill No","Item Name","Qty Returned","Unit Price",
                               "Return Amount","Reason"]
                for col, header in enumerate(headers_ret, 1):
                    cell = ws_ret.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                changed = True

            # Add Type column to ProfitLoss if missing
            if "ProfitLoss" in wb.sheetnames:
                ws_pl = wb["ProfitLoss"]
                if ws_pl.cell(1, 7).value != "Type":
                    ws_pl.cell(1, 7).value = "Type"
                    ws_pl.cell(1, 7).font = Font(bold=True, color="FFFFFF", size=11)
                    ws_pl.cell(1, 7).fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                    changed = True

            # Add missing settings
            ws_settings = wb["Settings"]
            existing_keys = set()
            for row in range(2, ws_settings.max_row + 1):
                k = ws_settings.cell(row, 1).value
                if k: existing_keys.add(k)

            new_settings = [("return_counter", 1), ("session_timeout_minutes", 10)]
            for key, val in new_settings:
                if key not in existing_keys:
                    ws_settings.append([key, val])
                    changed = True

            if changed:
                wb.save(self.filename)
        except Exception as e:
            print(f"Migration warning: {e}")

    # ─── Settings ───────────────────────────────
    def _get_setting(self, key, default=''):
        wb = load_workbook(self.filename)
        ws = wb["Settings"]
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, 1).value == key:
                val = ws.cell(row, 2).value
                return val if val is not None else default
        return default

    def _set_setting(self, key, value):
        wb = load_workbook(self.filename)
        ws = wb["Settings"]
        found = False
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, 1).value == key:
                ws.cell(row, 2).value = value
                found = True
                break
        if not found:
            ws.append([key, value])
        wb.save(self.filename)

    def get_store_settings(self):
        return {
            'store_name':      self._get_setting('store_name', 'My Store'),
            'store_tagline':   self._get_setting('store_tagline', 'Quality You Can Trust'),
            'store_address':   self._get_setting('store_address', ''),
            'store_phone':     self._get_setting('store_phone', ''),
            'store_email':     self._get_setting('store_email', ''),
            'store_website':   self._get_setting('store_website', ''),
            'currency_symbol': self._get_setting('currency_symbol', '₹'),
            'profit_margin':   float(self._get_setting('profit_margin', 30)),
        }

    def update_store_settings(self, settings_dict):
        for key, value in settings_dict.items():
            self._set_setting(key, value)

    def get_profit_margin(self):
        return float(self._get_setting('profit_margin', 30))

    def set_profit_margin(self, value):
        self._set_setting('profit_margin', value)

    def get_currency_symbol(self):
        return self._get_setting('currency_symbol', '₹')

    def get_session_timeout(self):
        """Returns session timeout in minutes (default 10)."""
        try:
            return int(self._get_setting('session_timeout_minutes', 10))
        except:
            return 10

    def set_session_timeout(self, minutes):
        """Admin sets staff session timeout in minutes."""
        self._set_setting('session_timeout_minutes', int(minutes))

    # ─── Customer ID ────────────────────────────
    def get_or_create_customer_id(self, customer_name):
        if not customer_name:
            return "CUST-0000"
        wb = load_workbook(self.filename)
        ws_bills = wb["Bills"]
        customer_ids = {}
        for row in range(2, ws_bills.max_row + 1):
            name = ws_bills.cell(row, 3).value
            cid  = ws_bills.cell(row, 4).value
            if name and cid:
                customer_ids[name.strip().lower()] = cid
        if customer_name.strip().lower() in customer_ids:
            return customer_ids[customer_name.strip().lower()]
        ws_settings = wb["Settings"]
        counter = 1000
        for row in range(2, ws_settings.max_row + 1):
            if ws_settings.cell(row, 1).value == "customer_counter":
                counter = int(ws_settings.cell(row, 2).value) + 1
                ws_settings.cell(row, 2).value = counter
                break
        wb.save(self.filename)
        return f"CUST-{counter}"

    # ─── Bill Number ────────────────────────────
    def generate_bill_no(self):
        wb = load_workbook(self.filename)
        ws = wb["Settings"]
        counter = 1
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, 1).value == "bill_counter":
                counter = int(ws.cell(row, 2).value)
                ws.cell(row, 2).value = counter + 1
                break
        wb.save(self.filename)
        date = datetime.now().strftime("%Y%m%d")
        return f"BILL-{date}-{counter:04d}"

    # ─── Save Bill ──────────────────────────────
    def save_bill_data(self, customer_name, customer_id, bill_no, items,
                       subtotal, discount_total, grand_total, qr_file, pdf_file):
        wb = load_workbook(self.filename)
        ws_bills = wb["Bills"]
        ws_pl    = wb["ProfitLoss"]

        sr_no    = ws_bills.max_row
        date_str = datetime.now().strftime("%d-%m-%Y")
        row_data = [
            sr_no, date_str, customer_name, customer_id, bill_no,
            len(items), subtotal, discount_total, grand_total, qr_file, pdf_file
        ]
        ws_bills.append(row_data)
        last_row = ws_bills.max_row
        for col in range(1, 12):
            cell = ws_bills.cell(last_row, col)
            cell.alignment = Alignment(horizontal="left" if col in [3,5,10,11] else "right")
            if last_row % 2 == 0:
                cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

        profit_margin = self.get_profit_margin() / 100.0
        cost   = subtotal * (1 - profit_margin)
        profit = (subtotal - cost) - discount_total

        ws_pl.append([bill_no, date_str, subtotal, discount_total, grand_total, profit, "SALE"])
        last_row_pl = ws_pl.max_row
        for col in range(1, 8):
            cell = ws_pl.cell(last_row_pl, col)
            cell.alignment = Alignment(horizontal="right" if col > 2 else "left")
            if last_row_pl % 2 == 0:
                cell.fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")

        wb.save(self.filename)
        return True

    # ─── Delete Single Bill ─────────────────────
    def delete_bill(self, bill_no):
        """Delete a single bill from Bills and ProfitLoss sheets by bill_no."""
        try:
            wb = load_workbook(self.filename)
            ws_bills = wb["Bills"]
            ws_pl    = wb["ProfitLoss"]

            # Delete from Bills sheet
            bills_deleted = False
            for row in range(ws_bills.max_row, 1, -1):
                if str(ws_bills.cell(row, 5).value) == str(bill_no):
                    ws_bills.delete_rows(row)
                    bills_deleted = True
                    break

            # Delete from ProfitLoss sheet
            for row in range(ws_pl.max_row, 1, -1):
                if str(ws_pl.cell(row, 1).value) == str(bill_no):
                    ws_pl.delete_rows(row)
                    break

            wb.save(self.filename)
            return bills_deleted
        except Exception as e:
            print(f"Error deleting bill: {e}")
            return False

    # ─── Returns ────────────────────────────────
    def save_return(self, customer_name, customer_id, original_bill_no,
                    items_returned, total_return_amount, reason=''):
        """Save a return transaction. Reduces sales/profit in ProfitLoss."""
        try:
            wb = load_workbook(self.filename)
            ws_ret = wb["Returns"]
            ws_pl  = wb["ProfitLoss"]

            # Generate return bill number
            ws_settings = wb["Settings"]
            ret_counter = 1
            for row in range(2, ws_settings.max_row + 1):
                if ws_settings.cell(row, 1).value == "return_counter":
                    ret_counter = int(ws_settings.cell(row, 2).value)
                    ws_settings.cell(row, 2).value = ret_counter + 1
                    break

            date_str   = datetime.now().strftime("%d-%m-%Y")
            return_no  = f"RET-{datetime.now().strftime('%Y%m%d')}-{ret_counter:04d}"
            sr_base    = max(1, ws_ret.max_row)

            for idx, item in enumerate(items_returned):
                row_data = [
                    sr_base + idx,
                    date_str,
                    customer_name,
                    customer_id,
                    original_bill_no,
                    return_no,
                    item.get('name', ''),
                    item.get('quantity', 1),
                    item.get('unit_price', 0),
                    item.get('return_amount', 0),
                    reason
                ]
                ws_ret.append(row_data)
                last_row = ws_ret.max_row
                for col in range(1, 12):
                    cell = ws_ret.cell(last_row, col)
                    cell.alignment = Alignment(horizontal="left" if col in [3,5,6,7,11] else "right")
                    if last_row % 2 == 0:
                        cell.fill = PatternFill(start_color="FDECEA", end_color="FDECEA", fill_type="solid")

            # Record negative entry in ProfitLoss to reduce totals
            profit_margin  = self.get_profit_margin() / 100.0
            return_profit  = -(total_return_amount * profit_margin)
            pl_data = [
                return_no, date_str,
                -total_return_amount,   # negative subtotal (reduces sales)
                0,                      # discount
                -total_return_amount,   # negative total
                return_profit,          # negative profit
                "RETURN"
            ]
            ws_pl.append(pl_data)
            last_pl = ws_pl.max_row
            for col in range(1, 8):
                cell = ws_pl.cell(last_pl, col)
                cell.alignment = Alignment(horizontal="right" if col > 2 else "left")
                cell.fill = PatternFill(start_color="FDECEA", end_color="FDECEA", fill_type="solid")

            wb.save(self.filename)
            return return_no
        except Exception as e:
            print(f"Error saving return: {e}")
            return None

    def get_all_returns(self):
        """Get all return transactions."""
        try:
            wb   = load_workbook(self.filename, data_only=True)
            if "Returns" not in wb.sheetnames:
                return []
            ws   = wb["Returns"]
            seen = set()
            returns = []
            for row in range(2, ws.max_row + 1):
                ret_no = ws.cell(row, 6).value
                if not ret_no:
                    continue
                if ret_no not in seen:
                    seen.add(ret_no)
                    returns.append({
                        'sr_no':            ws.cell(row, 1).value,
                        'date':             ws.cell(row, 2).value,
                        'customer_name':    ws.cell(row, 3).value,
                        'customer_id':      ws.cell(row, 4).value,
                        'original_bill_no': ws.cell(row, 5).value,
                        'return_no':        ret_no,
                        'item_name':        ws.cell(row, 7).value,
                        'return_amount':    float(ws.cell(row, 10).value or 0),
                        'reason':           ws.cell(row, 11).value or '',
                    })
            return list(reversed(returns))
        except Exception as e:
            print(f"Error fetching returns: {e}")
            return []

    # ─── Retrieve Bills ─────────────────────────
    def get_recent_bills(self, limit=100000):
        wb    = load_workbook(self.filename, data_only=True)
        ws    = wb["Bills"]
        bills = []
        for row in range(2, ws.max_row + 1):
            sr = ws.cell(row, 1).value
            if sr:
                bills.append({
                    'sr_no':         sr,
                    'date':          ws.cell(row, 2).value,
                    'customer_name': ws.cell(row, 3).value,
                    'customer_id':   ws.cell(row, 4).value,
                    'bill_no':       ws.cell(row, 5).value,
                    'items_count':   ws.cell(row, 6).value,
                    'subtotal':      float(ws.cell(row, 7).value or 0),
                    'discount':      float(ws.cell(row, 8).value or 0),
                    'total':         float(ws.cell(row, 9).value or 0),
                    'qr_file':       ws.cell(row, 10).value,
                    'pdf_file':      ws.cell(row, 11).value,
                })
        return list(reversed(bills))

    def get_all_bills(self):
        return self.get_recent_bills()

    # ─── Summary ────────────────────────────────
    def generate_summary_report(self):
        wb     = load_workbook(self.filename, data_only=True)
        ws_pl  = wb["ProfitLoss"]
        total_sales = 0; total_discounts = 0; total_profit = 0; total_returns = 0
        for row in range(2, ws_pl.max_row + 1):
            row_type = ws_pl.cell(row, 7).value or "SALE"
            subtotal = ws_pl.cell(row, 3).value or 0
            discount = ws_pl.cell(row, 4).value or 0
            profit   = ws_pl.cell(row, 6).value or 0
            if row_type == "RETURN":
                total_returns += abs(subtotal)
            total_sales     += subtotal
            total_discounts += discount
            total_profit    += profit

        ws_bills   = wb["Bills"]
        total_bills = max(0, ws_bills.max_row - 1)

        ret_count = 0
        if "Returns" in wb.sheetnames:
            ws_ret = wb["Returns"]
            seen   = set()
            for row in range(2, ws_ret.max_row + 1):
                rn = ws_ret.cell(row, 6).value
                if rn: seen.add(rn)
            ret_count = len(seen)

        return {
            'total_bills':     total_bills,
            'total_sales':     max(0, total_sales),
            'total_discounts': total_discounts,
            'total_profit':    total_profit,
            'total_returns':   total_returns,
            'total_returns_count': ret_count,
            'profit_margin':   self.get_profit_margin(),
            'session_timeout': self.get_session_timeout(),
        }

    def get_excel_file(self):
        return self.filename

    # ─── Reset ──────────────────────────────────
    def erase_all_bills(self):
        try:
            wb = load_workbook(self.filename)
            for sheet_name in ["Bills", "ProfitLoss", "Returns"]:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    if ws.max_row > 1:
                        ws.delete_rows(2, ws.max_row - 1)
            wb.save(self.filename)
            return True
        except Exception as e:
            print(f"Error erasing bills: {e}")
            return False

    def factory_reset(self):
        try:
            wb = load_workbook(self.filename)
            for sheet_name in ["Bills", "ProfitLoss", "Returns"]:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    if ws.max_row > 1:
                        ws.delete_rows(2, ws.max_row - 1)
            ws_settings = wb["Settings"]
            defaults = {
                "profit_margin": 30.0, "customer_counter": 1000,
                "bill_counter": 1, "return_counter": 1,
                "store_name": "My Store", "store_tagline": "Quality You Can Trust",
                "store_address": "123 Main Street, City - 000000",
                "store_phone": "+91 98765 43210", "store_email": "contact@mystore.com",
                "store_website": "www.mystore.com", "currency_symbol": "₹",
                "session_timeout_minutes": 10
            }
            for row in range(2, ws_settings.max_row + 1):
                key = ws_settings.cell(row, 1).value
                if key in defaults:
                    ws_settings.cell(row, 2).value = defaults[key]
            wb.save(self.filename)
            return True
        except Exception as e:
            print(f"Error during factory reset: {e}")
            return False
